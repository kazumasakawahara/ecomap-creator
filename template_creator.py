#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
エコマップExcelテンプレート作成スクリプト
===========================================
10シート構成の空のテンプレートを作成します。

使い方:
    python template_creator.py

出力:
    templates/template.xlsx
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date


def create_header_style():
    """ヘッダー行のスタイルを作成"""
    return {
        'font': Font(bold=True, color='FFFFFF', size=11),
        'fill': PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center'),
        'border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    }


def apply_header_style(cell):
    """セルにヘッダースタイルを適用"""
    style = create_header_style()
    cell.font = style['font']
    cell.fill = style['fill']
    cell.alignment = style['alignment']
    cell.border = style['border']


def set_column_widths(ws, widths):
    """列幅を設定"""
    for col_num, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width


def create_person_info_sheet(wb):
    """1. 本人情報シートを作成"""
    ws = wb.create_sheet("本人情報")
    
    # ヘッダー行
    headers = [
        "氏名（必須）", "生年月日（必須）\n※YYYY-MM-DD形式", "年齢\n※自動計算", 
        "性別（必須）\n※男/女/その他", "住所", "郵便番号", 
        "電話番号", "緊急連絡先", "備考"
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        apply_header_style(cell)
        ws.row_dimensions[1].height = 30
    
    # 列幅設定
    set_column_widths(ws, [15, 18, 8, 15, 30, 12, 15, 15, 30])
    
    # サンプル行（入力例として1行追加）
    ws.cell(row=2, column=1, value="例：山田太郎")
    ws.cell(row=2, column=2, value="例：1990-04-01")
    ws.cell(row=2, column=3, value="※自動")
    ws.cell(row=2, column=4, value="例：男")
    
    # 入力欄（3行目以降を空白に）
    for row in range(3, 6):
        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col, value="")
    
    return ws


def create_family_info_sheet(wb):
    """2. 家族情報シートを作成"""
    ws = wb.create_sheet("家族情報")
    
    headers = [
        "氏名（必須）", "続柄（必須）", "生年月日\n※YYYY-MM-DD形式", 
        "年齢\n※自動計算", "性別", "同居（必須）\n※〇/×", 
        "主介護者\n※〇/×", "住所", "電話番号", "備考"
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        apply_header_style(cell)
        ws.row_dimensions[1].height = 30
    
    set_column_widths(ws, [15, 12, 18, 8, 10, 12, 12, 30, 15, 30])
    
    # サンプル行
    ws.cell(row=2, column=1, value="例：山田花子")
    ws.cell(row=2, column=2, value="例：母")
    ws.cell(row=2, column=3, value="例：1965-06-15")
    ws.cell(row=2, column=4, value="※自動")
    ws.cell(row=2, column=5, value="例：女")
    ws.cell(row=2, column=6, value="例：〇")
    ws.cell(row=2, column=7, value="例：〇")
    
    return ws


def create_notebook_info_sheet(wb):
    """3. 手帳情報シートを作成"""
    ws = wb.create_sheet("手帳情報")
    
    headers = [
        "手帳種別（必須）\n※療育手帳/精神保健福祉手帳/身体障害者手帳", 
        "等級・判定（必須）", "手帳番号", "交付日（必須）\n※YYYY-MM-DD形式",
        "有効期限\n※YYYY-MM-DD形式", "交付自治体（必須）", 
        "状態（必須）\n※有効/期限切れ/更新済み", "備考"
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        apply_header_style(cell)
        ws.row_dimensions[1].height = 45
    
    set_column_widths(ws, [25, 15, 15, 18, 18, 15, 18, 30])
    
    # サンプル行
    ws.cell(row=2, column=1, value="例：療育手帳")
    ws.cell(row=2, column=2, value="例：B1")
    ws.cell(row=2, column=3, value="例：第123456号")
    ws.cell(row=2, column=4, value="例：2020-04-01")
    ws.cell(row=2, column=5, value="例：2025-03-31")
    ws.cell(row=2, column=6, value="例：北九州市")
    ws.cell(row=2, column=7, value="例：有効")
    
    return ws


def create_support_level_sheet(wb):
    """4. 支援区分情報シートを作成"""
    ws = wb.create_sheet("支援区分情報")
    
    headers = [
        "支援区分（必須）\n※0-6", "決定日（必須）\n※YYYY-MM-DD形式",
        "有効期限\n※YYYY-MM-DD形式", "決定自治体（必須）",
        "認定調査員", "状態（必須）\n※現在/期限切れ", "備考"
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        apply_header_style(cell)
        ws.row_dimensions[1].height = 30
    
    set_column_widths(ws, [15, 18, 18, 15, 20, 18, 30])
    
    # サンプル行
    ws.cell(row=2, column=1, value="例：3")
    ws.cell(row=2, column=2, value="例：2023-07-01")
    ws.cell(row=2, column=3, value="例：2026-06-30")
    ws.cell(row=2, column=4, value="例：北九州市")
    ws.cell(row=2, column=5, value="例：佐藤一郎")
    ws.cell(row=2, column=6, value="例：現在")
    
    return ws


def create_diagnosis_sheet(wb):
    """5. 診断情報シートを作成"""
    ws = wb.create_sheet("診断情報")
    
    headers = [
        "診断名（必須）", "ICD-10コード\n※任意", "診断日\n※YYYY-MM-DD形式",
        "診断医", "医療機関", "状態（必須）\n※継続/寛解/治癒", "備考"
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        apply_header_style(cell)
        ws.row_dimensions[1].height = 30
    
    set_column_widths(ws, [25, 15, 18, 15, 25, 18, 30])
    
    # サンプル行
    ws.cell(row=2, column=1, value="例：自閉スペクトラム症")
    ws.cell(row=2, column=2, value="例：F84.0")
    ws.cell(row=2, column=3, value="例：2018-03-15")
    ws.cell(row=2, column=4, value="例：田中医師")
    ws.cell(row=2, column=5, value="例：小倉病院")
    ws.cell(row=2, column=6, value="例：継続")
    
    return ws


def create_guardian_sheet(wb):
    """6. 成年後見情報シートを作成"""
    ws = wb.create_sheet("成年後見情報")
    
    headers = [
        "後見人氏名（必須）", "後見類型（必須）\n※後見/保佐/補助/任意後見",
        "後見人種別（必須）\n※親族/専門職", "専門職種\n※弁護士/司法書士等",
        "開始日（必須）\n※YYYY-MM-DD形式", "権限範囲", "連絡先", "備考"
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        apply_header_style(cell)
        ws.row_dimensions[1].height = 45
    
    set_column_widths(ws, [18, 20, 18, 20, 18, 25, 15, 30])
    
    # サンプル行
    ws.cell(row=2, column=1, value="例：鈴木次郎")
    ws.cell(row=2, column=2, value="例：後見")
    ws.cell(row=2, column=3, value="例：専門職")
    ws.cell(row=2, column=4, value="例：司法書士")
    ws.cell(row=2, column=5, value="例：2022-09-01")
    ws.cell(row=2, column=6, value="例：財産管理・身上監護")
    
    return ws


def create_consultation_support_sheet(wb):
    """7. 相談支援情報シートを作成"""
    ws = wb.create_sheet("相談支援情報")
    
    headers = [
        "事業所名（必須）", "事業所番号\n※10桁", 
        "支援種別（必須）\n※特定相談支援/一般相談支援",
        "担当専門員（必須）", "住所", "電話番号",
        "契約日\n※YYYY-MM-DD形式", "備考"
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        apply_header_style(cell)
        ws.row_dimensions[1].height = 30
    
    set_column_widths(ws, [25, 15, 22, 15, 30, 15, 18, 30])
    
    # サンプル行
    ws.cell(row=2, column=1, value="例：あおぞら相談支援センター")
    ws.cell(row=2, column=2, value="例：4012345678")
    ws.cell(row=2, column=3, value="例：特定相談支援")
    ws.cell(row=2, column=4, value="例：高橋三郎")
    ws.cell(row=2, column=5, value="例：北九州市小倉北区○○町1-1-1")
    ws.cell(row=2, column=6, value="例：093-123-4567")
    ws.cell(row=2, column=7, value="例：2021-04-01")
    
    return ws


def create_service_plan_sheet(wb):
    """8. サービス等利用計画シートを作成"""
    ws = wb.create_sheet("サービス等利用計画")
    
    headers = [
        "計画番号", "作成日（必須）\n※YYYY-MM-DD形式",
        "前回モニタリング日\n※YYYY-MM-DD形式", 
        "次回モニタリング予定日\n※YYYY-MM-DD形式",
        "状態（必須）\n※有効/期限切れ/見直し中", "備考"
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        apply_header_style(cell)
        ws.row_dimensions[1].height = 45
    
    set_column_widths(ws, [15, 18, 22, 25, 22, 30])
    
    # サンプル行
    ws.cell(row=2, column=1, value="例：2023-001")
    ws.cell(row=2, column=2, value="例：2023-04-01")
    ws.cell(row=2, column=3, value="例：2024-04-01")
    ws.cell(row=2, column=4, value="例：2025-04-01")
    ws.cell(row=2, column=5, value="例：有効")
    
    return ws


def create_service_usage_sheet(wb):
    """9. サービス利用情報シートを作成"""
    ws = wb.create_sheet("サービス利用情報")
    
    headers = [
        "サービス種別（必須）", "事業所名（必須）", "事業所番号\n※10桁",
        "サービス管理責任者", "住所", "電話番号",
        "契約日（必須）\n※YYYY-MM-DD形式", "利用頻度", "利用曜日",
        "状態（必須）\n※契約中/体験中/契約終了", "備考"
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        apply_header_style(cell)
        ws.row_dimensions[1].height = 30
    
    set_column_widths(ws, [20, 25, 15, 18, 30, 15, 18, 12, 15, 20, 30])
    
    # サンプル行
    ws.cell(row=2, column=1, value="例：就労継続支援B型")
    ws.cell(row=2, column=2, value="例：すみれ作業所")
    ws.cell(row=2, column=3, value="例：4012345679")
    ws.cell(row=2, column=4, value="例：木村四郎")
    ws.cell(row=2, column=5, value="例：北九州市八幡西区△△町2-2-2")
    ws.cell(row=2, column=6, value="例：093-234-5678")
    ws.cell(row=2, column=7, value="例：2022-06-01")
    ws.cell(row=2, column=8, value="例：週5日")
    ws.cell(row=2, column=9, value="例：月-金")
    ws.cell(row=2, column=10, value="例：契約中")
    
    return ws


def create_medical_institution_sheet(wb):
    """10. 医療機関情報シートを作成"""
    ws = wb.create_sheet("医療機関情報")
    
    headers = [
        "医療機関名（必須）", "診療科（必須）", "担当医",
        "主治医\n※〇/×", "住所", "電話番号",
        "通院開始日\n※YYYY-MM-DD形式", "通院頻度", "治療内容",
        "処方薬\n※薬剤名・用量・用法", "備考"
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        apply_header_style(cell)
        ws.row_dimensions[1].height = 30
    
    set_column_widths(ws, [25, 15, 15, 12, 30, 15, 18, 12, 20, 30, 30])
    
    # サンプル行
    ws.cell(row=2, column=1, value="例：小倉メンタルクリニック")
    ws.cell(row=2, column=2, value="例：精神科")
    ws.cell(row=2, column=3, value="例：山口医師")
    ws.cell(row=2, column=4, value="例：〇")
    ws.cell(row=2, column=5, value="例：北九州市小倉南区□□町3-3-3")
    ws.cell(row=2, column=6, value="例：093-345-6789")
    ws.cell(row=2, column=7, value="例：2019-01-15")
    ws.cell(row=2, column=8, value="例：月1回")
    ws.cell(row=2, column=9, value="例：精神療法・薬物療法")
    ws.cell(row=2, column=10, value="例：リスペリドン2mg 朝夕")
    
    return ws


def main():
    """メイン関数"""
    print("📝 エコマップExcelテンプレートを作成中...")
    
    # ワークブック作成
    wb = Workbook()
    
    # デフォルトシート削除
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # 各シートを作成
    print("  → 1. 本人情報シート作成中...")
    create_person_info_sheet(wb)
    
    print("  → 2. 家族情報シート作成中...")
    create_family_info_sheet(wb)
    
    print("  → 3. 手帳情報シート作成中...")
    create_notebook_info_sheet(wb)
    
    print("  → 4. 支援区分情報シート作成中...")
    create_support_level_sheet(wb)
    
    print("  → 5. 診断情報シート作成中...")
    create_diagnosis_sheet(wb)
    
    print("  → 6. 成年後見情報シート作成中...")
    create_guardian_sheet(wb)
    
    print("  → 7. 相談支援情報シート作成中...")
    create_consultation_support_sheet(wb)
    
    print("  → 8. サービス等利用計画シート作成中...")
    create_service_plan_sheet(wb)
    
    print("  → 9. サービス利用情報シート作成中...")
    create_service_usage_sheet(wb)
    
    print("  → 10. 医療機関情報シート作成中...")
    create_medical_institution_sheet(wb)
    
    # ファイル保存
    output_path = "templates/template.xlsx"
    wb.save(output_path)
    
    print(f"\n✅ テンプレート作成完了！")
    print(f"📁 保存先: {output_path}")
    print(f"📊 シート数: {len(wb.sheetnames)}枚")
    print("\n作成されたシート:")
    for i, sheet_name in enumerate(wb.sheetnames, 1):
        print(f"  {i:2d}. {sheet_name}")


if __name__ == "__main__":
    main()
