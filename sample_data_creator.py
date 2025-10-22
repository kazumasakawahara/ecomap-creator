#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
エコマップサンプルデータ作成スクリプト
=====================================
北九州市および隣接市町村の架空データを3パターン作成します。

使い方:
    python sample_data_creator.py

出力:
    samples/sample_case_01.xlsx - ケース1: 若年層・家族同居
    samples/sample_case_02.xlsx - ケース2: 中年層・グループホーム
    samples/sample_case_03.xlsx - ケース3: 高齢層・成年後見あり
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date, datetime
import sys


def create_header_style():
    """ヘッダー行のスタイルを作成"""
    return {
        'font': Font(bold=True, color='FFFFFF', size=11),
        'fill': PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center'),
        'border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    }


def apply_header_style(cell):
    """セルにヘッダースタイルを適用"""
    style = create_header_style()
    cell.font = style['font']
    cell.fill = style['fill']
    cell.alignment = style['alignment']
    cell.border = style['border']


def set_column_widths(ws, widths):
    """列幅を設定"""
    for col_num, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width


def add_person_info(ws, person_data):
    """本人情報シートにデータを追加"""
    headers = [
        "氏名（必須）", "生年月日（必須）\n※YYYY-MM-DD形式", "年齢\n※自動計算", 
        "性別（必須）\n※男/女/その他", "住所", "郵便番号", 
        "電話番号", "緊急連絡先", "備考"
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        apply_header_style(cell)
        ws.row_dimensions[1].height = 30
    
    set_column_widths(ws, [15, 18, 8, 15, 30, 12, 15, 15, 30])
    
    # データ行
    birth_date = datetime.strptime(person_data['birth_date'], '%Y-%m-%d').date()
    age = (date.today() - birth_date).days // 365
    
    ws.cell(row=2, column=1, value=person_data['name'])
    ws.cell(row=2, column=2, value=person_data['birth_date'])
    ws.cell(row=2, column=3, value=age)
    ws.cell(row=2, column=4, value=person_data['gender'])
    ws.cell(row=2, column=5, value=person_data['address'])
    ws.cell(row=2, column=6, value=person_data['postal_code'])
    ws.cell(row=2, column=7, value=person_data['phone'])
    ws.cell(row=2, column=8, value=person_data['emergency_contact'])
    ws.cell(row=2, column=9, value=person_data.get('note', ''))


def add_family_info(ws, family_data_list):
    """家族情報シートにデータを追加"""
    headers = [
        "氏名（必須）", "続柄（必須）", "生年月日\n※YYYY-MM-DD形式", 
        "年齢\n※自動計算", "性別", "同居（必須）\n※〇/×", 
        "主介護者\n※〇/×", "住所", "電話番号", "備考"
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        apply_header_style(cell)
        ws.row_dimensions[1].height = 30
    
    set_column_widths(ws, [15, 12, 18, 8, 10, 12, 12, 30, 15, 30])
    
    # データ行
    for row_num, family in enumerate(family_data_list, 2):
        birth_date = datetime.strptime(family['birth_date'], '%Y-%m-%d').date()
        age = (date.today() - birth_date).days // 365
        
        ws.cell(row=row_num, column=1, value=family['name'])
        ws.cell(row=row_num, column=2, value=family['relationship'])
        ws.cell(row=row_num, column=3, value=family['birth_date'])
        ws.cell(row=row_num, column=4, value=age)
        ws.cell(row=row_num, column=5, value=family['gender'])
        ws.cell(row=row_num, column=6, value=family['living_together'])
        ws.cell(row=row_num, column=7, value=family['primary_caregiver'])
        ws.cell(row=row_num, column=8, value=family.get('address', ''))
        ws.cell(row=row_num, column=9, value=family.get('phone', ''))
        ws.cell(row=row_num, column=10, value=family.get('note', ''))


def add_notebook_info(ws, notebook_data_list):
    """手帳情報シートにデータを追加"""
    headers = [
        "手帳種別（必須）\n※療育手帳/精神保健福祉手帳/身体障害者手帳", 
        "等級・判定（必須）", "手帳番号", "交付日（必須）\n※YYYY-MM-DD形式",
        "有効期限\n※YYYY-MM-DD形式", "交付自治体（必須）", 
        "状態（必須）\n※有効/期限切れ/更新済み", "備考"
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        apply_header_style(cell)
        ws.row_dimensions[1].height = 45
    
    set_column_widths(ws, [25, 15, 15, 18, 18, 15, 18, 30])
    
    # データ行
    for row_num, notebook in enumerate(notebook_data_list, 2):
        ws.cell(row=row_num, column=1, value=notebook['type'])
        ws.cell(row=row_num, column=2, value=notebook['grade'])
        ws.cell(row=row_num, column=3, value=notebook['number'])
        ws.cell(row=row_num, column=4, value=notebook['issue_date'])
        ws.cell(row=row_num, column=5, value=notebook.get('expiry_date', ''))
        ws.cell(row=row_num, column=6, value=notebook['issuing_authority'])
        ws.cell(row=row_num, column=7, value=notebook['status'])
        ws.cell(row=row_num, column=8, value=notebook.get('note', ''))


def add_support_level_info(ws, support_level_data_list):
    """支援区分情報シートにデータを追加"""
    headers = [
        "支援区分（必須）\n※0-6", "決定日（必須）\n※YYYY-MM-DD形式",
        "有効期限\n※YYYY-MM-DD形式", "決定自治体（必須）",
        "認定調査員", "状態（必須）\n※現在/期限切れ", "備考"
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        apply_header_style(cell)
        ws.row_dimensions[1].height = 30
    
    set_column_widths(ws, [15, 18, 18, 15, 20, 18, 30])
    
    # データ行
    for row_num, support in enumerate(support_level_data_list, 2):
        ws.cell(row=row_num, column=1, value=support['level'])
        ws.cell(row=row_num, column=2, value=support['decision_date'])
        ws.cell(row=row_num, column=3, value=support.get('expiry_date', ''))
        ws.cell(row=row_num, column=4, value=support['deciding_authority'])
        ws.cell(row=row_num, column=5, value=support.get('investigator', ''))
        ws.cell(row=row_num, column=6, value=support['status'])
        ws.cell(row=row_num, column=7, value=support.get('note', ''))


def add_diagnosis_info(ws, diagnosis_data_list):
    """診断情報シートにデータを追加"""
    headers = [
        "診断名（必須）", "ICD-10コード\n※任意", "診断日\n※YYYY-MM-DD形式",
        "診断医", "医療機関", "状態（必須）\n※継続/寛解/治癒", "備考"
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        apply_header_style(cell)
        ws.row_dimensions[1].height = 30
    
    set_column_widths(ws, [25, 15, 18, 15, 25, 18, 30])
    
    # データ行
    for row_num, diagnosis in enumerate(diagnosis_data_list, 2):
        ws.cell(row=row_num, column=1, value=diagnosis['name'])
        ws.cell(row=row_num, column=2, value=diagnosis.get('icd10_code', ''))
        ws.cell(row=row_num, column=3, value=diagnosis.get('diagnosis_date', ''))
        ws.cell(row=row_num, column=4, value=diagnosis.get('doctor', ''))
        ws.cell(row=row_num, column=5, value=diagnosis.get('institution', ''))
        ws.cell(row=row_num, column=6, value=diagnosis['status'])
        ws.cell(row=row_num, column=7, value=diagnosis.get('note', ''))


def add_guardian_info(ws, guardian_data_list):
    """成年後見情報シートにデータを追加"""
    headers = [
        "後見人氏名（必須）", "後見類型（必須）\n※後見/保佐/補助/任意後見",
        "後見人種別（必須）\n※親族/専門職", "専門職種\n※弁護士/司法書士等",
        "開始日（必須）\n※YYYY-MM-DD形式", "権限範囲", "連絡先", "備考"
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        apply_header_style(cell)
        ws.row_dimensions[1].height = 45
    
    set_column_widths(ws, [18, 20, 18, 20, 18, 25, 15, 30])
    
    # データ行
    for row_num, guardian in enumerate(guardian_data_list, 2):
        ws.cell(row=row_num, column=1, value=guardian['name'])
        ws.cell(row=row_num, column=2, value=guardian['type'])
        ws.cell(row=row_num, column=3, value=guardian['category'])
        ws.cell(row=row_num, column=4, value=guardian.get('profession', ''))
        ws.cell(row=row_num, column=5, value=guardian['start_date'])
        ws.cell(row=row_num, column=6, value=guardian.get('authority', ''))
        ws.cell(row=row_num, column=7, value=guardian.get('contact', ''))
        ws.cell(row=row_num, column=8, value=guardian.get('note', ''))


def add_consultation_support_info(ws, consultation_data_list):
    """相談支援情報シートにデータを追加"""
    headers = [
        "事業所名（必須）", "事業所番号\n※10桁", 
        "支援種別（必須）\n※特定相談支援/一般相談支援",
        "担当専門員（必須）", "住所", "電話番号",
        "契約日\n※YYYY-MM-DD形式", "備考"
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        apply_header_style(cell)
        ws.row_dimensions[1].height = 30
    
    set_column_widths(ws, [25, 15, 22, 15, 30, 15, 18, 30])
    
    # データ行
    for row_num, consultation in enumerate(consultation_data_list, 2):
        ws.cell(row=row_num, column=1, value=consultation['office_name'])
        ws.cell(row=row_num, column=2, value=consultation.get('office_number', ''))
        ws.cell(row=row_num, column=3, value=consultation['support_type'])
        ws.cell(row=row_num, column=4, value=consultation['specialist'])
        ws.cell(row=row_num, column=5, value=consultation.get('address', ''))
        ws.cell(row=row_num, column=6, value=consultation.get('phone', ''))
        ws.cell(row=row_num, column=7, value=consultation.get('contract_date', ''))
        ws.cell(row=row_num, column=8, value=consultation.get('note', ''))


def add_service_plan_info(ws, plan_data_list):
    """サービス等利用計画シートにデータを追加"""
    headers = [
        "計画番号", "作成日（必須）\n※YYYY-MM-DD形式",
        "前回モニタリング日\n※YYYY-MM-DD形式", 
        "次回モニタリング予定日\n※YYYY-MM-DD形式",
        "状態（必須）\n※有効/期限切れ/見直し中", "備考"
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        apply_header_style(cell)
        ws.row_dimensions[1].height = 45
    
    set_column_widths(ws, [15, 18, 22, 25, 22, 30])
    
    # データ行
    for row_num, plan in enumerate(plan_data_list, 2):
        ws.cell(row=row_num, column=1, value=plan.get('plan_number', ''))
        ws.cell(row=row_num, column=2, value=plan['creation_date'])
        ws.cell(row=row_num, column=3, value=plan.get('last_monitoring_date', ''))
        ws.cell(row=row_num, column=4, value=plan.get('next_monitoring_date', ''))
        ws.cell(row=row_num, column=5, value=plan['status'])
        ws.cell(row=row_num, column=6, value=plan.get('note', ''))


def add_service_usage_info(ws, service_data_list):
    """サービス利用情報シートにデータを追加"""
    headers = [
        "サービス種別（必須）", "事業所名（必須）", "事業所番号\n※10桁",
        "サービス管理責任者", "住所", "電話番号",
        "契約日（必須）\n※YYYY-MM-DD形式", "利用頻度", "利用曜日",
        "状態（必須）\n※契約中/体験中/契約終了", "備考"
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        apply_header_style(cell)
        ws.row_dimensions[1].height = 30
    
    set_column_widths(ws, [20, 25, 15, 18, 30, 15, 18, 12, 15, 20, 30])
    
    # データ行
    for row_num, service in enumerate(service_data_list, 2):
        ws.cell(row=row_num, column=1, value=service['service_type'])
        ws.cell(row=row_num, column=2, value=service['office_name'])
        ws.cell(row=row_num, column=3, value=service.get('office_number', ''))
        ws.cell(row=row_num, column=4, value=service.get('service_manager', ''))
        ws.cell(row=row_num, column=5, value=service.get('address', ''))
        ws.cell(row=row_num, column=6, value=service.get('phone', ''))
        ws.cell(row=row_num, column=7, value=service['contract_date'])
        ws.cell(row=row_num, column=8, value=service.get('frequency', ''))
        ws.cell(row=row_num, column=9, value=service.get('days_of_week', ''))
        ws.cell(row=row_num, column=10, value=service['status'])
        ws.cell(row=row_num, column=11, value=service.get('note', ''))


def add_medical_institution_info(ws, medical_data_list):
    """医療機関情報シートにデータを追加"""
    headers = [
        "医療機関名（必須）", "診療科（必須）", "担当医",
        "主治医\n※〇/×", "住所", "電話番号",
        "通院開始日\n※YYYY-MM-DD形式", "通院頻度", "治療内容",
        "処方薬\n※薬剤名・用量・用法", "備考"
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        apply_header_style(cell)
        ws.row_dimensions[1].height = 30
    
    set_column_widths(ws, [25, 15, 15, 12, 30, 15, 18, 12, 20, 30, 30])
    
    # データ行
    for row_num, medical in enumerate(medical_data_list, 2):
        ws.cell(row=row_num, column=1, value=medical['institution_name'])
        ws.cell(row=row_num, column=2, value=medical['department'])
        ws.cell(row=row_num, column=3, value=medical.get('doctor', ''))
        ws.cell(row=row_num, column=4, value=medical.get('primary_doctor', ''))
        ws.cell(row=row_num, column=5, value=medical.get('address', ''))
        ws.cell(row=row_num, column=6, value=medical.get('phone', ''))
        ws.cell(row=row_num, column=7, value=medical.get('start_date', ''))
        ws.cell(row=row_num, column=8, value=medical.get('frequency', ''))
        ws.cell(row=row_num, column=9, value=medical.get('treatment', ''))
        ws.cell(row=row_num, column=10, value=medical.get('medication', ''))
        ws.cell(row=row_num, column=11, value=medical.get('note', ''))


# ========================================
# サンプルケース1: 若年層・家族同居
# ========================================
def create_case_01():
    """ケース1: 若年層・家族同居のサンプルデータ"""
    wb = Workbook()
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # 1. 本人情報
    ws_person = wb.create_sheet("本人情報")
    person_data = {
        'name': '佐藤健太',
        'birth_date': '2001-07-15',
        'gender': '男',
        'address': '福岡県北九州市小倉北区西港町1-1-1',
        'postal_code': '803-0801',
        'phone': '090-1234-5678',
        'emergency_contact': '093-111-2222（母）',
        'note': '自閉スペクトラム症。コミュニケーションに配慮が必要'
    }
    add_person_info(ws_person, person_data)
    
    # 2. 家族情報
    ws_family = wb.create_sheet("家族情報")
    family_data = [
        {
            'name': '佐藤美智子',
            'relationship': '母',
            'birth_date': '1970-03-20',
            'gender': '女',
            'living_together': '〇',
            'primary_caregiver': '〇',
            'address': '福岡県北九州市小倉北区西港町1-1-1',
            'phone': '093-111-2222',
            'note': '主たる介護者。パート勤務'
        },
        {
            'name': '佐藤正樹',
            'relationship': '父',
            'birth_date': '1968-11-05',
            'gender': '男',
            'living_together': '〇',
            'primary_caregiver': '×',
            'address': '福岡県北九州市小倉北区西港町1-1-1',
            'phone': '090-2345-6789',
            'note': '会社員。平日は帰宅が遅い'
        },
        {
            'name': '佐藤由美',
            'relationship': '妹',
            'birth_date': '2004-09-12',
            'gender': '女',
            'living_together': '〇',
            'primary_caregiver': '×',
            'phone': '090-3456-7890',
            'note': '大学生。兄のサポートに協力的'
        }
    ]
    add_family_info(ws_family, family_data)
    
    # 3. 手帳情報
    ws_notebook = wb.create_sheet("手帳情報")
    notebook_data = [
        {
            'type': '療育手帳',
            'grade': 'B1',
            'number': '第401-123456号',
            'issue_date': '2020-04-01',
            'expiry_date': '2030-03-31',
            'issuing_authority': '北九州市',
            'status': '有効',
            'note': '令和2年交付'
        }
    ]
    add_notebook_info(ws_notebook, notebook_data)
    
    # 4. 支援区分情報
    ws_support = wb.create_sheet("支援区分情報")
    support_data = [
        {
            'level': '3',
            'decision_date': '2023-07-01',
            'expiry_date': '2026-06-30',
            'deciding_authority': '北九州市',
            'investigator': '田中調査員',
            'status': '現在',
            'note': ''
        }
    ]
    add_support_level_info(ws_support, support_data)
    
    # 5. 診断情報
    ws_diagnosis = wb.create_sheet("診断情報")
    diagnosis_data = [
        {
            'name': '自閉スペクトラム症',
            'icd10_code': 'F84.0',
            'diagnosis_date': '2008-05-10',
            'doctor': '山口医師',
            'institution': '小倉療育センター',
            'status': '継続',
            'note': '幼少期より診断'
        }
    ]
    add_diagnosis_info(ws_diagnosis, diagnosis_data)
    
    # 6. 成年後見情報（なし）
    ws_guardian = wb.create_sheet("成年後見情報")
    add_guardian_info(ws_guardian, [])
    
    # 7. 相談支援情報
    ws_consultation = wb.create_sheet("相談支援情報")
    consultation_data = [
        {
            'office_name': 'あおぞら相談支援センター',
            'office_number': '4012345678',
            'support_type': '特定相談支援',
            'specialist': '高橋三郎',
            'address': '福岡県北九州市小倉北区浅野1-1-1',
            'phone': '093-123-4567',
            'contract_date': '2021-04-01',
            'note': ''
        }
    ]
    add_consultation_support_info(ws_consultation, consultation_data)
    
    # 8. サービス等利用計画
    ws_plan = wb.create_sheet("サービス等利用計画")
    plan_data = [
        {
            'plan_number': '2024-001',
            'creation_date': '2024-04-01',
            'last_monitoring_date': '2024-10-01',
            'next_monitoring_date': '2025-04-01',
            'status': '有効',
            'note': '就労継続支援B型の利用を計画'
        }
    ]
    add_service_plan_info(ws_plan, plan_data)
    
    # 9. サービス利用情報
    ws_service = wb.create_sheet("サービス利用情報")
    service_data = [
        {
            'service_type': '就労継続支援B型',
            'office_name': 'すみれ作業所',
            'office_number': '4012345679',
            'service_manager': '木村四郎',
            'address': '福岡県北九州市八幡西区黒崎1-1-1',
            'phone': '093-234-5678',
            'contract_date': '2022-06-01',
            'frequency': '週5日',
            'days_of_week': '月-金',
            'status': '契約中',
            'note': 'パン製造作業に従事'
        }
    ]
    add_service_usage_info(ws_service, service_data)
    
    # 10. 医療機関情報
    ws_medical = wb.create_sheet("医療機関情報")
    medical_data = [
        {
            'institution_name': '小倉メンタルクリニック',
            'department': '精神科',
            'doctor': '山口医師',
            'primary_doctor': '〇',
            'address': '福岡県北九州市小倉南区葉山町1-1-1',
            'phone': '093-345-6789',
            'start_date': '2019-01-15',
            'frequency': '月1回',
            'treatment': '精神療法・薬物療法',
            'medication': 'リスペリドン2mg 朝夕、バルプロ酸200mg 朝昼夕',
            'note': '定期通院継続中'
        }
    ]
    add_medical_institution_info(ws_medical, medical_data)
    
    return wb


# ========================================
# サンプルケース2: 中年層・グループホーム
# ========================================
def create_case_02():
    """ケース2: 中年層・グループホーム入居のサンプルデータ"""
    wb = Workbook()
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # 1. 本人情報
    ws_person = wb.create_sheet("本人情報")
    person_data = {
        'name': '田中花子',
        'birth_date': '1980-12-03',
        'gender': '女',
        'address': '福岡県北九州市門司区大里本町1-1-1（グループホームひまわり）',
        'postal_code': '800-0001',
        'phone': '080-9876-5432',
        'emergency_contact': '093-222-3333（GH職員）',
        'note': '統合失調症。グループホーム入居中'
    }
    add_person_info(ws_person, person_data)
    
    # 2. 家族情報
    ws_family = wb.create_sheet("家族情報")
    family_data = [
        {
            'name': '田中良子',
            'relationship': '母',
            'birth_date': '1955-08-25',
            'gender': '女',
            'living_together': '×',
            'primary_caregiver': '×',
            'address': '福岡県北九州市門司区東港町2-2-2',
            'phone': '093-222-3334',
            'note': '高齢。定期的に面会'
        }
    ]
    add_family_info(ws_family, family_data)
    
    # 3. 手帳情報
    ws_notebook = wb.create_sheet("手帳情報")
    notebook_data = [
        {
            'type': '精神保健福祉手帳',
            'grade': '2級',
            'number': '第401-234567号',
            'issue_date': '2023-04-01',
            'expiry_date': '2025-03-31',
            'issuing_authority': '福岡県',
            'status': '有効',
            'note': '令和5年度更新'
        }
    ]
    add_notebook_info(ws_notebook, notebook_data)
    
    # 4. 支援区分情報
    ws_support = wb.create_sheet("支援区分情報")
    support_data = [
        {
            'level': '4',
            'decision_date': '2022-09-01',
            'expiry_date': '2025-08-31',
            'deciding_authority': '北九州市',
            'investigator': '鈴木調査員',
            'status': '現在',
            'note': ''
        }
    ]
    add_support_level_info(ws_support, support_data)
    
    # 5. 診断情報
    ws_diagnosis = wb.create_sheet("診断情報")
    diagnosis_data = [
        {
            'name': '統合失調症',
            'icd10_code': 'F20.0',
            'diagnosis_date': '1998-11-20',
            'doctor': '中村医師',
            'institution': '門司病院',
            'status': '継続',
            'note': '病状は安定'
        }
    ]
    add_diagnosis_info(ws_diagnosis, diagnosis_data)
    
    # 6. 成年後見情報（なし）
    ws_guardian = wb.create_sheet("成年後見情報")
    add_guardian_info(ws_guardian, [])
    
    # 7. 相談支援情報
    ws_consultation = wb.create_sheet("相談支援情報")
    consultation_data = [
        {
            'office_name': 'そよかぜ相談支援事業所',
            'office_number': '4012345680',
            'support_type': '特定相談支援',
            'specialist': '伊藤美香',
            'address': '福岡県北九州市門司区大里本町3-3-3',
            'phone': '093-456-7890',
            'contract_date': '2020-10-01',
            'note': ''
        }
    ]
    add_consultation_support_info(ws_consultation, consultation_data)
    
    # 8. サービス等利用計画
    ws_plan = wb.create_sheet("サービス等利用計画")
    plan_data = [
        {
            'plan_number': '2024-002',
            'creation_date': '2024-10-01',
            'last_monitoring_date': '2024-10-01',
            'next_monitoring_date': '2025-10-01',
            'status': '有効',
            'note': 'GH・就労継続支援A型の利用継続'
        }
    ]
    add_service_plan_info(ws_plan, plan_data)
    
    # 9. サービス利用情報
    ws_service = wb.create_sheet("サービス利用情報")
    service_data = [
        {
            'service_type': '共同生活援助（グループホーム）',
            'office_name': 'グループホームひまわり',
            'office_number': '4012345681',
            'service_manager': '林太郎',
            'address': '福岡県北九州市門司区大里本町1-1-1',
            'phone': '093-567-8901',
            'contract_date': '2018-04-01',
            'frequency': '常時',
            'days_of_week': '毎日',
            'status': '契約中',
            'note': '個室利用。夜間支援あり'
        },
        {
            'service_type': '就労継続支援A型',
            'office_name': 'ワークスペースもじ',
            'office_number': '4012345682',
            'service_manager': '清水次郎',
            'address': '福岡県北九州市門司区港町2-2-2',
            'phone': '093-678-9012',
            'contract_date': '2021-07-01',
            'frequency': '週5日',
            'days_of_week': '月-金',
            'status': '契約中',
            'note': 'データ入力業務に従事'
        }
    ]
    add_service_usage_info(ws_service, service_data)
    
    # 10. 医療機関情報
    ws_medical = wb.create_sheet("医療機関情報")
    medical_data = [
        {
            'institution_name': '門司病院',
            'department': '精神科',
            'doctor': '中村医師',
            'primary_doctor': '〇',
            'address': '福岡県北九州市門司区大里東1-1-1',
            'phone': '093-789-0123',
            'start_date': '1998-11-20',
            'frequency': '月1回',
            'treatment': '精神療法・薬物療法・デイケア',
            'medication': 'オランザピン10mg 就寝前、フルボキサミン50mg 朝夕',
            'note': '長期通院継続中。病状安定'
        }
    ]
    add_medical_institution_info(ws_medical, medical_data)
    
    return wb


# ========================================
# サンプルケース3: 高齢層・成年後見あり
# ========================================
def create_case_03():
    """ケース3: 高齢層・成年後見制度利用のサンプルデータ"""
    wb = Workbook()
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # 1. 本人情報
    ws_person = wb.create_sheet("本人情報")
    person_data = {
        'name': '鈴木正雄',
        'birth_date': '1962-05-18',
        'gender': '男',
        'address': '山口県下関市竹崎町1-1-1',
        'postal_code': '750-0003',
        'phone': '083-111-2222',
        'emergency_contact': '083-222-3333（後見人）',
        'note': '知的障害。成年後見制度利用中'
    }
    add_person_info(ws_person, person_data)
    
    # 2. 家族情報（なし - 身寄りなし）
    ws_family = wb.create_sheet("家族情報")
    add_family_info(ws_family, [])
    
    # 3. 手帳情報
    ws_notebook = wb.create_sheet("手帳情報")
    notebook_data = [
        {
            'type': '療育手帳',
            'grade': 'A',
            'number': '第35-123456号',
            'issue_date': '2015-05-01',
            'expiry_date': '2025-04-30',
            'issuing_authority': '山口県',
            'status': '有効',
            'note': '平成27年交付'
        },
        {
            'type': '身体障害者手帳',
            'grade': '4級（下肢機能障害）',
            'number': '第35-234567号',
            'issue_date': '2019-03-15',
            'expiry_date': '',
            'issuing_authority': '下関市',
            'status': '有効',
            'note': '再交付不要'
        }
    ]
    add_notebook_info(ws_notebook, notebook_data)
    
    # 4. 支援区分情報
    ws_support = wb.create_sheet("支援区分情報")
    support_data = [
        {
            'level': '5',
            'decision_date': '2023-01-01',
            'expiry_date': '2026-12-31',
            'deciding_authority': '下関市',
            'investigator': '佐々木調査員',
            'status': '現在',
            'note': '高齢化に伴い区分変更'
        }
    ]
    add_support_level_info(ws_support, support_data)
    
    # 5. 診断情報
    ws_diagnosis = wb.create_sheet("診断情報")
    diagnosis_data = [
        {
            'name': '中等度知的障害',
            'icd10_code': 'F71',
            'diagnosis_date': '1970-04-01',
            'doctor': '前田医師',
            'institution': '下関市立病院',
            'status': '継続',
            'note': '幼少期より診断'
        },
        {
            'name': '変形性膝関節症',
            'icd10_code': 'M17',
            'diagnosis_date': '2018-06-10',
            'doctor': '加藤医師',
            'institution': '下関整形外科クリニック',
            'status': '継続',
            'note': '歩行困難あり'
        }
    ]
    add_diagnosis_info(ws_diagnosis, diagnosis_data)
    
    # 6. 成年後見情報
    ws_guardian = wb.create_sheet("成年後見情報")
    guardian_data = [
        {
            'name': '渡辺浩二',
            'type': '後見',
            'category': '専門職',
            'profession': '司法書士',
            'start_date': '2020-09-01',
            'authority': '財産管理・身上監護',
            'contact': '083-222-3333',
            'note': '下関市成年後見支援センター経由で選任'
        }
    ]
    add_guardian_info(ws_guardian, guardian_data)
    
    # 7. 相談支援情報
    ws_consultation = wb.create_sheet("相談支援情報")
    consultation_data = [
        {
            'office_name': 'しものせき相談支援事業所',
            'office_number': '3512345678',
            'support_type': '特定相談支援',
            'specialist': '吉田五郎',
            'address': '山口県下関市竹崎町2-2-2',
            'phone': '083-123-4567',
            'contract_date': '2019-04-01',
            'note': ''
        }
    ]
    add_consultation_support_info(ws_consultation, consultation_data)
    
    # 8. サービス等利用計画
    ws_plan = wb.create_sheet("サービス等利用計画")
    plan_data = [
        {
            'plan_number': '2024-003',
            'creation_date': '2024-04-01',
            'last_monitoring_date': '2024-07-01',
            'next_monitoring_date': '2025-01-01',
            'status': '有効',
            'note': '生活介護・短期入所の利用継続'
        }
    ]
    add_service_plan_info(ws_plan, plan_data)
    
    # 9. サービス利用情報
    ws_service = wb.create_sheet("サービス利用情報")
    service_data = [
        {
            'service_type': '生活介護',
            'office_name': 'しものせき生活支援センター',
            'office_number': '3512345679',
            'service_manager': '松本六郎',
            'address': '山口県下関市東大和町1-1-1',
            'phone': '083-234-5678',
            'contract_date': '2017-04-01',
            'frequency': '週5日',
            'days_of_week': '月-金',
            'status': '契約中',
            'note': '日中活動の場として利用。入浴介助あり'
        },
        {
            'service_type': '短期入所（ショートステイ）',
            'office_name': 'ケアホームあさひ',
            'office_number': '3512345680',
            'service_manager': '中島七郎',
            'address': '山口県下関市新地町3-3-3',
            'phone': '083-345-6789',
            'contract_date': '2019-10-01',
            'frequency': '月1-2回',
            'days_of_week': '不定期',
            'status': '契約中',
            'note': '後見人不在時や体調不良時に利用'
        }
    ]
    add_service_usage_info(ws_service, service_data)
    
    # 10. 医療機関情報
    ws_medical = wb.create_sheet("医療機関情報")
    medical_data = [
        {
            'institution_name': '下関市立病院',
            'department': '精神科',
            'doctor': '前田医師',
            'primary_doctor': '〇',
            'address': '山口県下関市向洋町1-1-1',
            'phone': '083-456-7890',
            'start_date': '1970-04-01',
            'frequency': '2か月に1回',
            'treatment': '精神療法・定期診察',
            'medication': '',
            'note': '長期通院。高齢化に伴う身体面の配慮必要'
        },
        {
            'institution_name': '下関整形外科クリニック',
            'department': '整形外科',
            'doctor': '加藤医師',
            'primary_doctor': '×',
            'address': '山口県下関市南大坪町2-2-2',
            'phone': '083-567-8901',
            'start_date': '2018-06-10',
            'frequency': '月1回',
            'treatment': 'リハビリテーション・投薬',
            'medication': 'ロキソプロフェン60mg 頓服',
            'note': '膝関節痛の管理'
        }
    ]
    add_medical_institution_info(ws_medical, medical_data)
    
    return wb


def main():
    """メイン関数"""
    print("📝 エコマップサンプルデータを作成中...")
    
    # ケース1
    print("\n  → ケース1: 若年層・家族同居（佐藤健太さん）")
    wb1 = create_case_01()
    output_path1 = "samples/sample_case_01.xlsx"
    wb1.save(output_path1)
    print(f"     ✅ 保存完了: {output_path1}")
    
    # ケース2
    print("\n  → ケース2: 中年層・グループホーム（田中花子さん）")
    wb2 = create_case_02()
    output_path2 = "samples/sample_case_02.xlsx"
    wb2.save(output_path2)
    print(f"     ✅ 保存完了: {output_path2}")
    
    # ケース3
    print("\n  → ケース3: 高齢層・成年後見あり（鈴木正雄さん）")
    wb3 = create_case_03()
    output_path3 = "samples/sample_case_03.xlsx"
    wb3.save(output_path3)
    print(f"     ✅ 保存完了: {output_path3}")
    
    print("\n" + "="*60)
    print("🎉 サンプルデータ作成完了！")
    print("="*60)
    print("\n📊 作成されたサンプル:")
    print("  1. sample_case_01.xlsx - 佐藤健太さん（24歳・男性・自閉スペクトラム症）")
    print("     - 家族同居、療育手帳B1、支援区分3、就労継続支援B型利用")
    print("  2. sample_case_02.xlsx - 田中花子さん（44歳・女性・統合失調症）")
    print("     - グループホーム入居、精神保健福祉手帳2級、支援区分4")
    print("     - 就労継続支援A型利用")
    print("  3. sample_case_03.xlsx - 鈴木正雄さん（62歳・男性・知的障害）")
    print("     - 成年後見制度利用、療育手帳A・身体障害者手帳4級、支援区分5")
    print("     - 生活介護・短期入所利用")


if __name__ == "__main__":
    main()
