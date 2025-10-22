#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Excel読み込みモジュール

Excelファイルから各シートのデータを読み込みます。
"""

import openpyxl
from typing import Dict, List, Any, Optional


class ExcelReader:
    """Excel読み込みクラス"""
    
    def __init__(self, file_path: str):
        """
        初期化
        
        Args:
            file_path: Excelファイルパス
        """
        self.file_path = file_path
        self.workbook = None
    
    def load(self) -> Dict[str, Any]:
        """
        Excelファイルを読み込み、全データを返す
        
        Returns:
            全データの辞書
        """
        try:
            self.workbook = openpyxl.load_workbook(self.file_path, data_only=True)
            
            data = {
                "person": self.read_person_info(),
                "family": self.read_family_info(),
                "notebooks": self.read_notebooks_info(),
                "support_levels": self.read_support_levels_info(),
                "diagnoses": self.read_diagnoses_info(),
                "legal_guardians": self.read_legal_guardians_info(),
                "consultation_supports": self.read_consultation_supports_info(),
                "service_plans": self.read_service_plans_info(),
                "service_contracts": self.read_service_contracts_info(),
                "medical_institutions": self.read_medical_institutions_info(),
            }
            
            return data
            
        finally:
            if self.workbook:
                self.workbook.close()
    
    def read_person_info(self) -> Dict[str, Any]:
        """本人情報シートを読み込み"""
        sheet_name = "本人情報"
        if sheet_name not in self.workbook.sheetnames:
            return {}
        
        ws = self.workbook[sheet_name]
        
        # 2行目からデータを読み込む（1行目はヘッダー）
        data = {}
        
        # セル位置を指定して読み込み
        # A列: 項目名、B列: 値
        row = 2
        data["name"] = self._get_cell_value(ws, f"B{row}")
        row += 1
        data["birth_date"] = self._get_cell_value(ws, f"B{row}")
        row += 1
        data["gender"] = self._get_cell_value(ws, f"B{row}")
        row += 1
        data["address"] = self._get_cell_value(ws, f"B{row}")
        row += 1
        data["postal_code"] = self._get_cell_value(ws, f"B{row}")
        row += 1
        data["phone"] = self._get_cell_value(ws, f"B{row}")
        row += 1
        data["emergency_contact"] = self._get_cell_value(ws, f"B{row}")
        row += 1
        data["notes"] = self._get_cell_value(ws, f"B{row}")
        
        return data
    
    def read_family_info(self) -> List[Dict[str, Any]]:
        """家族情報シートを読み込み"""
        sheet_name = "家族情報"
        if sheet_name not in self.workbook.sheetnames:
            return []
        
        ws = self.workbook[sheet_name]
        
        # ヘッダー行（2行目）を読み込み
        headers = []
        for col in range(1, 10):  # A-I列
            header = ws.cell(2, col).value
            if header:
                headers.append(header)
        
        # データ行（3行目以降）を読み込み
        data_list = []
        for row in range(3, ws.max_row + 1):
            # 氏名が空の行はスキップ
            name = self._get_cell_value(ws, f"A{row}")
            if not name:
                continue
            
            data = {
                "name": name,
                "relation": self._get_cell_value(ws, f"B{row}"),
                "birth_date": self._get_cell_value(ws, f"C{row}"),
                "gender": self._get_cell_value(ws, f"D{row}"),
                "living_together": self._get_cell_value(ws, f"E{row}"),
                "primary_caregiver": self._get_cell_value(ws, f"F{row}"),
                "address": self._get_cell_value(ws, f"G{row}"),
                "phone": self._get_cell_value(ws, f"H{row}"),
                "notes": self._get_cell_value(ws, f"I{row}"),
            }
            data_list.append(data)
        
        return data_list
    
    def read_notebooks_info(self) -> List[Dict[str, Any]]:
        """手帳情報シートを読み込み"""
        sheet_name = "手帳情報"
        if sheet_name not in self.workbook.sheetnames:
            return []
        
        ws = self.workbook[sheet_name]
        
        data_list = []
        for row in range(3, ws.max_row + 1):
            # 手帳種別が空の行はスキップ
            notebook_type = self._get_cell_value(ws, f"A{row}")
            if not notebook_type:
                continue
            
            data = {
                "type": notebook_type,
                "grade": self._get_cell_value(ws, f"B{row}"),
                "number": self._get_cell_value(ws, f"C{row}"),
                "issue_date": self._get_cell_value(ws, f"D{row}"),
                "expiry_date": self._get_cell_value(ws, f"E{row}"),
                "issuing_authority": self._get_cell_value(ws, f"F{row}"),
                "status": self._get_cell_value(ws, f"G{row}"),
                "notes": self._get_cell_value(ws, f"H{row}"),
            }
            data_list.append(data)
        
        return data_list
    
    def read_support_levels_info(self) -> List[Dict[str, Any]]:
        """支援区分情報シートを読み込み"""
        sheet_name = "支援区分情報"
        if sheet_name not in self.workbook.sheetnames:
            return []
        
        ws = self.workbook[sheet_name]
        
        data_list = []
        for row in range(3, ws.max_row + 1):
            # 支援区分が空の行はスキップ
            level = self._get_cell_value(ws, f"A{row}")
            if level is None or level == "":
                continue
            
            data = {
                "level": level,
                "decision_date": self._get_cell_value(ws, f"B{row}"),
                "expiry_date": self._get_cell_value(ws, f"C{row}"),
                "deciding_authority": self._get_cell_value(ws, f"D{row}"),
                "assessor": self._get_cell_value(ws, f"E{row}"),
                "status": self._get_cell_value(ws, f"F{row}"),
                "notes": self._get_cell_value(ws, f"G{row}"),
            }
            data_list.append(data)
        
        return data_list
    
    def read_diagnoses_info(self) -> List[Dict[str, Any]]:
        """診断情報シートを読み込み"""
        sheet_name = "診断情報"
        if sheet_name not in self.workbook.sheetnames:
            return []
        
        ws = self.workbook[sheet_name]
        
        data_list = []
        for row in range(3, ws.max_row + 1):
            # 診断名が空の行はスキップ
            diagnosis_name = self._get_cell_value(ws, f"A{row}")
            if not diagnosis_name:
                continue
            
            data = {
                "name": diagnosis_name,
                "icd10_code": self._get_cell_value(ws, f"B{row}"),
                "diagnosis_date": self._get_cell_value(ws, f"C{row}"),
                "doctor": self._get_cell_value(ws, f"D{row}"),
                "institution": self._get_cell_value(ws, f"E{row}"),
                "status": self._get_cell_value(ws, f"F{row}"),
                "notes": self._get_cell_value(ws, f"G{row}"),
            }
            data_list.append(data)
        
        return data_list
    
    def read_legal_guardians_info(self) -> List[Dict[str, Any]]:
        """成年後見情報シートを読み込み"""
        sheet_name = "成年後見情報"
        if sheet_name not in self.workbook.sheetnames:
            return []
        
        ws = self.workbook[sheet_name]
        
        data_list = []
        for row in range(3, ws.max_row + 1):
            # 後見人氏名が空の行はスキップ
            guardian_name = self._get_cell_value(ws, f"A{row}")
            if not guardian_name:
                continue
            
            data = {
                "name": guardian_name,
                "type": self._get_cell_value(ws, f"B{row}"),
                "category": self._get_cell_value(ws, f"C{row}"),
                "profession": self._get_cell_value(ws, f"D{row}"),
                "start_date": self._get_cell_value(ws, f"E{row}"),
                "authority": self._get_cell_value(ws, f"F{row}"),
                "contact": self._get_cell_value(ws, f"G{row}"),
                "notes": self._get_cell_value(ws, f"H{row}"),
            }
            data_list.append(data)
        
        return data_list
    
    def read_consultation_supports_info(self) -> List[Dict[str, Any]]:
        """相談支援情報シートを読み込み"""
        sheet_name = "相談支援情報"
        if sheet_name not in self.workbook.sheetnames:
            return []
        
        ws = self.workbook[sheet_name]
        
        data_list = []
        for row in range(3, ws.max_row + 1):
            # 事業所名が空の行はスキップ
            office_name = self._get_cell_value(ws, f"A{row}")
            if not office_name:
                continue
            
            data = {
                "office_name": office_name,
                "office_number": self._get_cell_value(ws, f"B{row}"),
                "support_type": self._get_cell_value(ws, f"C{row}"),
                "specialist": self._get_cell_value(ws, f"D{row}"),
                "address": self._get_cell_value(ws, f"E{row}"),
                "phone": self._get_cell_value(ws, f"F{row}"),
                "contract_date": self._get_cell_value(ws, f"G{row}"),
                "notes": self._get_cell_value(ws, f"H{row}"),
            }
            data_list.append(data)
        
        return data_list
    
    def read_service_plans_info(self) -> List[Dict[str, Any]]:
        """サービス等利用計画シートを読み込み"""
        sheet_name = "サービス等利用計画"
        if sheet_name not in self.workbook.sheetnames:
            return []
        
        ws = self.workbook[sheet_name]
        
        data_list = []
        for row in range(3, ws.max_row + 1):
            # 作成日が空の行はスキップ
            creation_date = self._get_cell_value(ws, f"B{row}")
            if not creation_date:
                continue
            
            data = {
                "plan_number": self._get_cell_value(ws, f"A{row}"),
                "creation_date": creation_date,
                "last_monitoring_date": self._get_cell_value(ws, f"C{row}"),
                "next_monitoring_date": self._get_cell_value(ws, f"D{row}"),
                "status": self._get_cell_value(ws, f"E{row}"),
                "notes": self._get_cell_value(ws, f"F{row}"),
            }
            data_list.append(data)
        
        return data_list
    
    def read_service_contracts_info(self) -> List[Dict[str, Any]]:
        """サービス利用情報シートを読み込み"""
        sheet_name = "サービス利用情報"
        if sheet_name not in self.workbook.sheetnames:
            return []
        
        ws = self.workbook[sheet_name]
        
        data_list = []
        for row in range(3, ws.max_row + 1):
            # サービス種別が空の行はスキップ
            service_type = self._get_cell_value(ws, f"A{row}")
            if not service_type:
                continue
            
            data = {
                "service_type": service_type,
                "office_name": self._get_cell_value(ws, f"B{row}"),
                "office_number": self._get_cell_value(ws, f"C{row}"),
                "manager": self._get_cell_value(ws, f"D{row}"),
                "address": self._get_cell_value(ws, f"E{row}"),
                "phone": self._get_cell_value(ws, f"F{row}"),
                "contract_date": self._get_cell_value(ws, f"G{row}"),
                "frequency": self._get_cell_value(ws, f"H{row}"),
                "days": self._get_cell_value(ws, f"I{row}"),
                "status": self._get_cell_value(ws, f"J{row}"),
                "notes": self._get_cell_value(ws, f"K{row}"),
            }
            data_list.append(data)
        
        return data_list
    
    def read_medical_institutions_info(self) -> List[Dict[str, Any]]:
        """医療機関情報シートを読み込み"""
        sheet_name = "医療機関情報"
        if sheet_name not in self.workbook.sheetnames:
            return []
        
        ws = self.workbook[sheet_name]
        
        data_list = []
        for row in range(3, ws.max_row + 1):
            # 医療機関名が空の行はスキップ
            institution_name = self._get_cell_value(ws, f"A{row}")
            if not institution_name:
                continue
            
            data = {
                "name": institution_name,
                "department": self._get_cell_value(ws, f"B{row}"),
                "doctor": self._get_cell_value(ws, f"C{row}"),
                "primary_doctor": self._get_cell_value(ws, f"D{row}"),
                "address": self._get_cell_value(ws, f"E{row}"),
                "phone": self._get_cell_value(ws, f"F{row}"),
                "start_date": self._get_cell_value(ws, f"G{row}"),
                "frequency": self._get_cell_value(ws, f"H{row}"),
                "treatment": self._get_cell_value(ws, f"I{row}"),
                "medications": self._get_cell_value(ws, f"J{row}"),
                "notes": self._get_cell_value(ws, f"K{row}"),
            }
            data_list.append(data)
        
        return data_list
    
    def _get_cell_value(self, ws, cell_ref: str) -> Any:
        """
        セルの値を取得（型変換なし）
        
        Args:
            ws: ワークシート
            cell_ref: セル参照（例: "A1"）
            
        Returns:
            セルの値（Noneの場合は空文字列）
        """
        value = ws[cell_ref].value
        
        # Noneを空文字列に変換
        if value is None:
            return ""
        
        # 文字列の場合、前後の空白を削除
        if isinstance(value, str):
            return value.strip()
        
        return value


if __name__ == "__main__":
    import sys
    
    if len(sys.argv) < 2:
        print("使用方法: python excel_reader.py <Excelファイル>")
        sys.exit(1)
    
    file_path = sys.argv[1]
    
    print(f"=== ExcelReader テスト ===")
    print(f"ファイル: {file_path}\n")
    
    try:
        reader = ExcelReader(file_path)
        data = reader.load()
        
        print("【本人情報】")
        print(f"氏名: {data['person'].get('name')}")
        print(f"生年月日: {data['person'].get('birth_date')}")
        print(f"性別: {data['person'].get('gender')}")
        
        print(f"\n【家族情報】{len(data['family'])}件")
        for family in data['family']:
            print(f"  - {family.get('name')} ({family.get('relation')})")
        
        print(f"\n【手帳情報】{len(data['notebooks'])}件")
        for notebook in data['notebooks']:
            print(f"  - {notebook.get('type')} {notebook.get('grade')}")
        
        print(f"\n【支援区分情報】{len(data['support_levels'])}件")
        for level in data['support_levels']:
            print(f"  - 区分{level.get('level')}")
        
        print(f"\n【サービス利用情報】{len(data['service_contracts'])}件")
        for contract in data['service_contracts']:
            print(f"  - {contract.get('service_type')}: {contract.get('office_name')}")
        
    except Exception as e:
        print(f"エラー: {e}")
        import traceback
        traceback.print_exc()
